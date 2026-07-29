import socket
import time
import openpyxl
from openpyxl import Workbook
import os
import math
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button

# Config
ESP32_IP = '192.168.4.1'  # Default ESP32 Access Point IP
PORT = 5005
EXCEL_FILE = 'registo_bateria.xlsx'
MAX_PLOT_POINTS = 100
BATTERY_DIVIDER_RATIO = 4.84  # Voltage divider factor from battery_monitor.ino (4.84:1 ratio)

# Global state
current_throttle_pulse = 1000  # Default off/armed pulse (1000us)
last_sent_pulse = None
active_client_socket = None
low_voltage_cutoff_active = False

def send_throttle_command(pulse, force=False):
    """Sends a throttle PWM command (1000us - 2000us) to ESP32 ONLY if the value changed or if forced."""
    global current_throttle_pulse, last_sent_pulse, active_client_socket, low_voltage_cutoff_active
    
    # If LVC active, force pulse to 1000us
    if low_voltage_cutoff_active:
        pulse = 1000

    current_throttle_pulse = int(pulse)
    if active_client_socket and (current_throttle_pulse != last_sent_pulse or force):
        try:
            cmd = f"THROTTLE:{current_throttle_pulse}\n"
            active_client_socket.sendall(cmd.encode('utf-8'))
            last_sent_pulse = current_throttle_pulse
            print(f"[GUI] Sent ESC Throttle Command: {current_throttle_pulse} us")
        except Exception as e:
            print(f"[GUI Error] Failed to send throttle command: {e}")

def update_live_plot(fig, ax, line, text_info, timestamps, raw_adcs, throttle_slider=None):
    """Updates the live matplotlib graph downsampling points to at most MAX_PLOT_POINTS while keeping full range."""
    global low_voltage_cutoff_active
    N = len(timestamps)
    if N == 0:
        return
    
    if N <= MAX_PLOT_POINTS:
        plot_x = list(timestamps)
        plot_y = list(raw_adcs)
    else:
        step = math.ceil(N / MAX_PLOT_POINTS)
        plot_x = list(timestamps[::step])
        plot_y = list(raw_adcs[::step])
        # Guarantee that the latest data point (current state) is always included
        if plot_x[-1] != timestamps[-1]:
            plot_x.append(timestamps[-1])
            plot_y.append(raw_adcs[-1])

    line.set_data(plot_x, plot_y)
    
    # Recalculate plot limits
    ax.relim()
    ax.autoscale_view()
    
    latest_adc = raw_adcs[-1]
    latest_t = timestamps[-1]
    pin_v = (latest_adc * 3.3) / 4095.0
    est_v = pin_v * BATTERY_DIVIDER_RATIO

    # Low Voltage Cutoff safety check on raw ADC (approx <= 2500 counts / 2.0V pin voltage -> 12.0V battery)
    if latest_adc <= 2500.0 and latest_adc > 0.0:
        low_voltage_cutoff_active = True
        if throttle_slider and throttle_slider.val > 1000:
            throttle_slider.set_val(1000)

    step_val = 1 if N <= MAX_PLOT_POINTS else math.ceil(N / MAX_PLOT_POINTS)
    step_str = f"1 (100% res)" if step_val == 1 else f"{step_val} (Showing {len(plot_x)} / {N} pts)"
    
    if low_voltage_cutoff_active:
        text_info.set_text(
            f"CRITICAL: RAW SENSOR <= 2500 (LOW VOLTAGE)! THROTTLE BLOCKED FOR SAFETY!\nElapsed: {latest_t:.1f}s  |  Raw ADC: {latest_adc:.1f}  |  Pin: {pin_v:.2f}V  |  Est Batt: {est_v:.2f}V"
        )
        text_info.set_bbox(dict(boxstyle='round,pad=0.5', facecolor='#660000', alpha=0.9, edgecolor='#FF0000'))
    else:
        throttle_str = f"OFF (1000us)" if current_throttle_pulse == 1000 else f"ON ({current_throttle_pulse}us)"
        text_info.set_text(
            f"Elapsed: {latest_t:.1f}s  |  Raw ADC: {latest_adc:.1f}  |  Pin: {pin_v:.2f}V  |  Est Batt: {est_v:.2f}V  |  Step: {step_str}  |  Throttle: {throttle_str}"
        )
        text_info.set_bbox(dict(boxstyle='round,pad=0.5', facecolor='#1E1E1E', alpha=0.85, edgecolor='#00E5FF'))

def main():
    global active_client_socket, current_throttle_pulse, low_voltage_cutoff_active

    print("==================================================")
    print("       Wireless Raw Battery Sensor Data Logger     ")
    print("==================================================")
    print(f"Target ESP32 AP IP : {ESP32_IP}:{PORT}")
    print(f"Wi-Fi SSID          : ESP32_Battery_Monitor")
    print(f"Max Plot Points     : {MAX_PLOT_POINTS} (Downsampled live graph)")
    print(f"Excel Columns       : [Record Number, Elapsed Time (s), Raw Sensor (ADC 0-4095), Raw Pin Voltage (V), Estimated Voltage (V)]")
    print("==================================================\n")
    
    records = []
    timestamps = []
    raw_adcs = []
    start_time = None

    # Load existing data from Excel if file exists
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        # Ensure header format has 5 columns: Record Number, Elapsed Time (s), Raw Sensor (ADC 0-4095), Raw Pin Voltage (V), Estimated Voltage (V)
        ws.cell(row=1, column=1, value="Record Number")
        ws.cell(row=1, column=2, value="Elapsed Time (s)")
        ws.cell(row=1, column=3, value="Raw Sensor (ADC 0-4095)")
        ws.cell(row=1, column=4, value="Raw Pin Voltage (V)")
        ws.cell(row=1, column=5, value="Estimated Voltage (V)")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2 and row[0] is not None:
                try:
                    rec_id = int(row[0])
                    if len(row) >= 3 and row[2] is not None:
                        t_elapsed = float(row[1])
                        adc_val = float(row[2])
                    else:
                        t_elapsed = float((rec_id - 1) * 1.0)
                        adc_val = float(row[1])
                    records.append(rec_id)
                    timestamps.append(t_elapsed)
                    raw_adcs.append(adc_val)
                except (ValueError, TypeError):
                    pass
        record_number = records[-1] + 1 if records else 1
        if timestamps:
            # Continue elapsed timing from last recorded elapsed time
            start_time = time.time() - timestamps[-1]
        print(f"Loaded {len(records)} existing records from '{EXCEL_FILE}'.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw Sensor Log"
        ws.append(["Record Number", "Elapsed Time (s)", "Raw Sensor (ADC 0-4095)", "Raw Pin Voltage (V)", "Estimated Voltage (V)"])
        record_number = 1
        wb.save(EXCEL_FILE)
        print(f"Created new log file '{EXCEL_FILE}'.")

    # Initialize live Matplotlib figure with dark theme
    plt.style.use('dark_background')
    plt.ion()
    fig, ax = plt.subplots(figsize=(11, 6.2))
    fig.canvas.manager.set_window_title("Wireless Battery Monitor - Raw Sensor GUI")

    # Adjust layout padding so plot, vertical slider, and buttons fit cleanly
    fig.subplots_adjust(left=0.08, right=0.82, top=0.90, bottom=0.18)

    line, = ax.plot([], [], color='#00E5FF', linewidth=2, marker='o', markersize=3.5, label='Raw Sensor Value (ADC)')
    ax.set_title("Live Raw Battery Sensor Logger (Unconverted ADC Data)", fontsize=14, pad=12, fontweight='bold', color='#FFFFFF')
    ax.set_xlabel("Elapsed Time (s)", fontsize=11)
    ax.set_ylabel("Raw Sensor ADC Reading (0 - 4095)", fontsize=11)
    ax.grid(True, linestyle='--', alpha=0.35)
    ax.legend(loc='upper right')

    text_info = ax.text(0.02, 0.95, "Initializing plot...", transform=ax.transAxes, fontsize=9.5,
                        verticalalignment='top', bbox=dict(boxstyle='round,pad=0.5', facecolor='#1E1E1E', alpha=0.85, edgecolor='#00E5FF'))

    # Add Vertical Slider for Throttle Control (1000us to 1500us, step 50)
    ax_slider = fig.add_axes([0.87, 0.20, 0.035, 0.65])
    throttle_slider = Slider(
        ax=ax_slider,
        label='Throttle\n(us)',
        valmin=1000,
        valmax=1500,
        valinit=1000,
        valstep=50,
        orientation='vertical',
        color='#00E5FF',
        valfmt='%d'
    )
    throttle_slider.label.set_color('#FFFFFF')
    throttle_slider.label.set_fontsize(10)
    throttle_slider.label.set_fontweight('bold')
    throttle_slider.valtext.set_color('#00E5FF')
    throttle_slider.valtext.set_fontweight('bold')

    # Add Emergency Stop Button
    ax_btn_stop = fig.add_axes([0.32, 0.03, 0.30, 0.075])
    btn_stop = Button(ax_btn_stop, 'EMERGENCY STOP (1000us)', color='#551111', hovercolor='#991111')
    btn_stop.label.set_color('#FFDDDD')
    btn_stop.label.set_fontsize(10)
    btn_stop.label.set_fontweight('bold')

    def on_slider_change(val):
        if low_voltage_cutoff_active:
            if throttle_slider.val > 1000:
                throttle_slider.set_val(1000)
                return
        pulse = int(val)
        send_throttle_command(pulse)
        if timestamps:
            update_live_plot(fig, ax, line, text_info, timestamps, raw_adcs, throttle_slider=throttle_slider)

    def on_click_stop(event):
        throttle_slider.set_val(1000)

    throttle_slider.on_changed(on_slider_change)
    btn_stop.on_clicked(on_click_stop)

    # Initial plot update if existing records were loaded
    if timestamps:
        update_live_plot(fig, ax, line, text_info, timestamps, raw_adcs, throttle_slider=throttle_slider)

    while plt.fignum_exists(fig.number):
        try:
            print(f"Connecting to ESP32 at {ESP32_IP}:{PORT}...")
            client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            client_socket.settimeout(0.05)  # Non-blocking 50ms timeout for smooth GUI redrawing
            client_socket.connect((ESP32_IP, PORT))
            active_client_socket = client_socket
            print("Connected to ESP32!")
            
            # Send initial throttle command on connect (forced)
            send_throttle_command(current_throttle_pulse, force=True)
            print("Receiving raw sensor data...\n")
            
            buffer = ""
            while plt.fignum_exists(fig.number):
                try:
                    data = client_socket.recv(1024)
                    if not data:
                        print("Connection closed by ESP32.")
                        break
                    
                    buffer += data.decode('utf-8', errors='ignore')
                    while '\n' in buffer:
                        line_str, buffer = buffer.split('\n', 1)
                        line_str = line_str.strip()
                        if line_str:
                            try:
                                raw_adc = float(line_str)
                                now = time.time()
                                if start_time is None:
                                    start_time = now
                                elapsed_sec = round(now - start_time, 2)
                                pin_voltage = round((raw_adc * 3.3) / 4095.0, 3)
                                estimated_voltage = round(pin_voltage * BATTERY_DIVIDER_RATIO, 2)
                                
                                # 1. Save Data to Excel: [Record Number, Elapsed Time (s), Raw Sensor (ADC 0-4095), Raw Pin Voltage (V), Estimated Voltage (V)]
                                ws.append([record_number, elapsed_sec, raw_adc, pin_voltage, estimated_voltage])
                                wb.save(EXCEL_FILE)
                                
                                # 2. Add to in-memory buffers
                                records.append(record_number)
                                timestamps.append(elapsed_sec)
                                raw_adcs.append(raw_adc)
                                
                                print(f"Record #{record_number} | {elapsed_sec:.2f}s | Raw ADC: {raw_adc:.1f} ({pin_voltage:.2f}V pin | {estimated_voltage:.2f}V est) saved.")
                                record_number += 1
                                
                                # 3. Update plot
                                update_live_plot(fig, ax, line, text_info, timestamps, raw_adcs, throttle_slider=throttle_slider)
                                
                            except ValueError:
                                pass
                except socket.timeout:
                    # Normal timeout - allows continuous GUI redrawing & event processing
                    pass

                # Keep Matplotlib GUI responsive to slider dragging, window minimize, restore, click, and redraw events
                if plt.fignum_exists(fig.number):
                    fig.canvas.flush_events()
                    plt.pause(0.02)

        except (socket.error, socket.timeout) as e:
            print(f"Connection error: {e}. Retrying in 3 seconds...")
            retry_start = time.time()
            while time.time() - retry_start < 3.0 and plt.fignum_exists(fig.number):
                fig.canvas.flush_events()
                plt.pause(0.05)
        except KeyboardInterrupt:
            print("\nLogging stopped by user.")
            break
        finally:
            if active_client_socket:
                try:
                    active_client_socket.close()
                except Exception:
                    pass
                active_client_socket = None

    if 'wb' in locals():
        wb.save(EXCEL_FILE)
    
    print("Program exited.")

if __name__ == "__main__":
    main()
